import streamlit as st
import pandas as pd
import joblib
import pickle
import os
import io
import warnings

# Suppress noisy library warnings
warnings.filterwarnings("ignore", category=UserWarning, module="torch")
warnings.filterwarnings("ignore", category=UserWarning, module="sklearn")
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# -------------------------------------------------
# Page Config
# -------------------------------------------------
st.set_page_config(
    page_title="Premier League Predictor (No Odds)",
    layout="wide"
)

# Initialize Session State for history
if 'prediction_history' not in st.session_state:
    st.session_state['prediction_history'] = []

# -------------------------------------------------
# Models Configuration (NO XGBOOST)
# -------------------------------------------------
MODELS = {
    "Random Forest": {
        "model": "football_model_rf_no_odds.pkl",
        "encoding": "team_encoding_rf_no_odds.pkl",
        "history": "full_match_history_rf_no_odds.pkl"
    },
    "Logistic Regression": {
        "model": "football_model_no_odds.pkl",
        "encoding": "team_encoding_no_odds.pkl",
        "history": "full_match_history_no_odds.pkl"
    },
    "XGBoost (Pure 26)": {
        "model": "football_model_pure26_no_odd.pkl",
        "encoding": "team_encoding_pure26_no_odd.pkl",
        "history": "full_match_history_no_odds.pkl"
    }
}

# -------------------------------------------------
# Load Resources (Cached)
# -------------------------------------------------
@st.cache_resource(show_spinner=False)
def load_all_models():
    loaded_data = {}

    for name, config in MODELS.items():
        try:
            if not os.path.exists(config["model"]):
                st.warning(f"⚠️ {name}: Model file not found.")
                continue

            model = joblib.load(config["model"])

            with open(config["encoding"], "rb") as f:
                encoding = pickle.load(f)

            with open(config["history"], "rb") as f:
                history = pickle.load(f)

            loaded_data[name] = {
                "model": model,
                "encoding": encoding,
                "history": history
            }

        except Exception as e:
            st.error(f"❌ Error loading {name}: {e}")

    return loaded_data


model_registry = load_all_models()

# -------------------------------------------------
# Helper Functions
# -------------------------------------------------
def get_rolling_stats(team_name, df, is_home_team):
    team_matches = df[df["HomeTeam"] == team_name]
    if team_matches.empty:
        return None

    if "Date" in team_matches.columns:
        last_match = team_matches.sort_values("Date", ascending=False).iloc[0]
    else:
        last_match = team_matches.iloc[-1]

    cols = ["FTHG", "FTAG", "HS", "AS", "HST", "AST"]
    rolling_cols = [f"{c}_rolling" for c in cols]

    stats = last_match[rolling_cols].to_dict()

    if not is_home_team:
        stats = {f"Away_{k}": v for k, v in stats.items()}

    return stats

# -------------------------------------------------
# UI
# -------------------------------------------------
st.title("Premier League Predictor")
st.markdown("### Compare predictions (No Betting Odds)")

if not model_registry:
    st.error("No models loaded successfully. Please check your files.")
    st.stop()

# Team list
first_model_key = list(model_registry.keys())[0]
teams = sorted(model_registry[first_model_key]["encoding"].keys())

st.sidebar.header("Match Details")

col1, col2 = st.columns(2)
with col1:
    home_team = st.selectbox("Home Team", teams)
with col2:
    away_team = st.selectbox("Away Team", teams, index=1)

col3, col4 = st.columns(2)
with col3:
    match_date = st.date_input("Match Date")
with col4:
    match_time = st.time_input("Match Time")

st.sidebar.divider()
gameweek = st.sidebar.text_input("Gameweek", value="Gameweek19")

# -------------------------------------------------
# OCR & Performance Optimization
# -------------------------------------------------
@st.cache_resource(show_spinner=False)
def get_ocr_reader():
    import easyocr
    return easyocr.Reader(['en'], gpu=False)

@st.cache_data(show_spinner=False)
def extract_text_from_image(image_bytes):
    import numpy as np
    from PIL import Image
    import io
    
    img = Image.open(io.BytesIO(image_bytes))
    img_np = np.array(img)
    reader = get_ocr_reader()
    return reader.readtext(img_np, detail=1)

st.sidebar.divider()
st.sidebar.subheader("Automated Match Extraction")
uploaded_file = st.sidebar.file_uploader("Upload Fixture Image", type=["png", "jpg", "jpeg"])

if uploaded_file is not None:
    try:
        # Read file as bytes for caching
        image_bytes = uploaded_file.getvalue()
        
        st.sidebar.info("Processing image... (Cached for speed)")
        result = extract_text_from_image(image_bytes)
        
        # Join lines and let the review
        extracted_text = "\n".join([res[1] for res in result])
        st.sidebar.success("Successfully extracted text!")
        
# Coordinate-Based Spatial Parsing
        import difflib
        import re
        
        formatted_lines = []
        
        # 1. Team Mapping Helper (Expanded)
        def get_fuzzy_team(text):
            if not text: return None
            t = text.lower().replace("'", "").replace(".", "").strip()
            
            # Direct Mappings for OCR inconsistencies
            if any(x in t for x in ["bumnie", "buml", "burn", "bumi"]): return "Burnley"
            if any(x in t for x in ["nexo", "newcastle", "newc", "castle"]): return "Newcastle"
            if any(x in t for x in ["chers", "chelsea", "chels", "chcl", "chersed"]): return "Chelsea"
            if any(x in t for x in ["bourn", "boum", "bour", "bom", "nemouu", "mouth"]): return "Bournemouth"
            if any(x in t for x in ["nott", "forest", "notte", "ednnn"]): return "Nott'm Forest"
            if any(x in t for x in ["arsen", "aisen", "arsn", "ars", "arsenal"]): return "Arsenal"
            if any(x in t for x in ["aston", "vilo", "villa", "ast"]): return "Aston Villa"
            if any(x in t for x in ["manut", "honut", "man utd", "man united", "unite", "man u"]): return "Man United"
            if any(x in t for x in ["bright", "bgt", "bri"]): return "Brighton"
            if any(x in t for x in ["evert", "ever", "eve"]): return "Everton"
            if any(x in t for x in ["west ham", "wvest", "whu", "west"]): return "West Ham"
            if any(x in t for x in ["wolv", "wol"]): return "Wolves"
            if any(x in t for x in ["cryst", "palace", "cpfc"]): return "Crystal Palace"
            if any(x in t for x in ["fulh", "fuln", "ful"]): return "Fulham"
            if any(x in t for x in ["liver", "liv", "lfc"]): return "Liverpool"
            if any(x in t for x in ["leeds", "leed", "lee"]): return "Leeds"
            if any(x in t for x in ["brent", "bre"]): return "Brentford"
            if any(x in t for x in ["spurs", "tott", "sduis", "hotspur"]): return "Tottenham"
            if any(x in t for x in ["sunder", "sun"]): return "Sunderland"
            if any(x in t for x in ["city", "aty", "mci"]): return "Man City"
            if any(x in t for x in ["luton"]): return "Luton"
            if any(x in t for x in ["sheff", "utd"]): return "Sheffield Utd"
            
            matches = difflib.get_close_matches(text, teams, n=1, cutoff=0.55)
            return matches[0] if matches else None

        # 2. Extract items with coordinates
        # result: [ [[x1,y1],[x2,y2],[x3,y3],[x4,y4]], text, confidence ]
        raw_items = []
        for res in result:
            box, text, conf = res
            y_center = (box[0][1] + box[2][1]) / 2
            x_center = (box[0][0] + box[1][0]) / 2
            raw_items.append({'y': y_center, 'x': x_center, 'text': text})

        # 3. Sort by Y first
        raw_items.sort(key=lambda x: x['y'])

        # 4. Group into rows
        rows = []
        if raw_items:
            curr_row = [raw_items[0]]
            for item in raw_items[1:]:
                # Group if vertical centers are within 20 pixels
                if abs(item['y'] - curr_row[0]['y']) < 20:
                    curr_row.append(item)
                else:
                    rows.append(curr_row)
                    curr_row = [item]
            rows.append(curr_row)

        # 5. Process Rows (Date Headers vs Match Rows)
        
        # Get global info from user input match date to infer Year if needed
        base_year = match_date.year
        current_date_str = match_date.strftime("%Y-%m-%d") # Default
        
        months_map = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6, 
                      "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

        formatted_lines = []

        for row in rows:
            row.sort(key=lambda x: x['x']) # Sort left to right
            row_text_all = " ".join([i['text'] for i in row])
            
            # --- Check for Date Header ---
            # Regex: (Sat|Sun|...) <Day> <Month> e.g. "Sat 3 Jan"
            date_match = re.search(r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)", row_text_all, re.IGNORECASE)
            
            if date_match:
                # Found a new date header!
                d_day = int(date_match.group(2))
                d_month_str = date_match.group(3).capitalize()
                d_month = months_map.get(d_month_str, 1)
                
                # Infer year
                # If current month is Dec and finding Jan, add 1 year
                extracted_year = base_year
                if match_date.month == 12 and d_month == 1:
                    extracted_year += 1
                elif match_date.month == 1 and d_month == 12:
                    extracted_year -= 1 # Unlikely but possible
                
                current_date_str = f"{extracted_year}-{d_month:02d}-{d_day:02d}"
                continue # Skip this row as it is just a header
            
            # --- Check for Match Row ---
            # Needs a time: HH:MM
            time_item = None
            for it in row:
                if re.search(r"(\d{1,2})[\s.:]+(\d{2})", it['text']):
                    time_item = it
                    break
            
            if time_item:
                # Extract and Clean Time
                t_match = re.search(r"(\d{1,2})[\s.:]+(\d{2})", time_item['text'])
                if t_match:
                    h, m = int(t_match.group(1)), int(t_match.group(2))
                    
                    # Logic to snap times or fix OCR errors
                    if 20 <= m <= 55: m = 30
                    elif m > 55 or m < 10: 
                        if m > 55: h = (h + 1) % 24
                        m = 0
                    elif 10 <= m <= 25: m = 15
                    
                    clean_time = f"{h:02d}:{m:02d}"
                    
                    # Find Home and Away teams based on position relative to Time
                    home_candidates = [i for i in row if i['x'] < time_item['x'] - 20] # 20px buffer
                    away_candidates = [i for i in row if i['x'] > time_item['x'] + 20]
                    
                    # Join text for better fuzzy matching (e.g. "Man" "City" -> "Man City")
                    h_text = " ".join([hc['text'] for hc in home_candidates])
                    a_text = " ".join([ac['text'] for ac in away_candidates])
                    
                    h_team = get_fuzzy_team(h_text)
                    a_team = get_fuzzy_team(a_text)
                    
                    if h_team and a_team and h_team != a_team:
                        formatted_lines.append(f"{h_team}, {a_team}, {current_date_str}, {clean_time}")

        if formatted_lines:
            st.session_state['bulk_input_value'] = "\n".join(formatted_lines)
            st.sidebar.success(f"Extracted {len(formatted_lines)} matches!")
        else:
            st.sidebar.warning("No matches extracted. Try a clearer image.")
            st.session_state['bulk_input_value'] = extracted_text

    except Exception as e:
        st.sidebar.error(f"Error processing image: {e}")

st.sidebar.divider()
st.sidebar.subheader("Bulk Prediction")
# Add a way to sync the gameweek
gw_val = st.session_state.get('bulk_gw', "Gameweek19")
gameweek = st.sidebar.text_input("Gameweek", value=gw_val, key="gw_input")

bulk_input = st.sidebar.text_area(
    "Paste matches (Format: Home, Away, YYYY-MM-DD, HH:MM)",
    value=st.session_state.get('bulk_input_value', ""),
    height=200,
    help="One match per line. Use commas to separate Home, Away, Date, and Time."
)

if st.sidebar.button("Process Bulk Predictions"):
    if not bulk_input.strip():
        st.sidebar.warning("Please enter match data.")
    else:
        # Use regex to find all matches in the input text (handles single line or multiline)
        # Pattern: Team A, Team B, YYYY-MM-DD, HH:MM
        # We allow loosely for teams (alphanum, spaces, quotes)
        import re
        pattern = r"([^,]+),\s*([^,]+),\s*(\d{4}-\d{2}-\d{2}),\s*(\d{1,2}:\d{2})"
        
        matches_found = re.findall(pattern, bulk_input)
        
        processed_count = 0
        
        if not matches_found:
             # Fallback to old splitting if regex fails completely (e.g. strict commas but weird format)
             st.sidebar.error("No valid matches found. Please check the format: 'Home, Away, YYYY-MM-DD, HH:MM'")
        
        for match in matches_found:
            try:
                h_team, a_team, d_str, t_str = [m.strip() for m in match]
                
                # Basic validation against team list
                if h_team not in teams or a_team not in teams:
                    st.sidebar.error(f"Team not found: {h_team} or {a_team}")
                    continue
                
                if h_team == a_team:
                    st.sidebar.error(f"Same teams: {line}")
                    continue

                # Run predictions for this match across all models
                match_dt = pd.to_datetime(f"{d_str} {t_str}")
                day_code = match_dt.weekday()
                hour = match_dt.hour
                
                current_match_results = []
                
                for model_name, assets in model_registry.items():
                    input_data = {
                        "home_code": assets["encoding"].get(h_team, -1),
                        "away_code": assets["encoding"].get(a_team, -1),
                        "hour": hour,
                        "day_code": day_code
                    }

                    home_stats = get_rolling_stats(h_team, assets["history"], True)
                    away_stats = get_rolling_stats(a_team, assets["history"], False)

                    if not home_stats or not away_stats:
                        continue

                    input_data.update(home_stats)
                    input_data.update(away_stats)

                    df_input = pd.DataFrame([input_data])
                    prediction = assets["model"].predict(df_input)[0]
                    result_code = {2: "H", 1: "D", 0: "A"}[prediction]
                    
                    current_match_results.append({
                        "Gameweek": gameweek,
                        "Date": match_dt.strftime("%d/%m/%Y"),
                        "Time": match_dt.strftime("%H:%M"),
                        "HomeTeam": h_team,
                        "AwayTeam": a_team,
                        "Model": model_name.lower().replace(" (pure 26)", ""),
                        "Result": result_code
                    })
                
                if current_match_results:
                    # Use current length of history to determine order
                    # (Divided by model count if we want per-match order)
                    order_val = len(st.session_state['prediction_history'])
                    for res in current_match_results:
                        res["Order"] = order_val
                    
                    st.session_state['prediction_history'].extend(current_match_results)
                    processed_count += 1
                    
            except Exception as e:
                st.sidebar.error(f"Error processing '{line}': {e}")
        
        if processed_count > 0:
            st.sidebar.success(f"Processed {processed_count} matches!")
            st.rerun()

# -------------------------------------------------
# Prediction Logic
# -------------------------------------------------
if st.button("Run Predictions", type="primary"):
    if home_team == away_team:
        st.error("Home and Away teams cannot be the same!")
        st.stop()

    st.divider()
    cols = st.columns(len(model_registry))

    # Temporary list for current run (all models)
    current_match_results = []
    
    for col, (model_name, assets) in zip(cols, model_registry.items()):
        with col:
            st.subheader(model_name)

            day_code = match_date.weekday()
            hour = match_time.hour

            input_data = {
                "home_code": assets["encoding"].get(home_team, -1),
                "away_code": assets["encoding"].get(away_team, -1),
                "hour": hour,
                "day_code": day_code
            }

            home_stats = get_rolling_stats(home_team, assets["history"], True)
            away_stats = get_rolling_stats(away_team, assets["history"], False)

            if not home_stats or not away_stats:
                st.warning("Insufficient history data.")
                continue

            input_data.update(home_stats)
            input_data.update(away_stats)

            try:
                df = pd.DataFrame([input_data])

                prediction = assets["model"].predict(df)[0]
                probs = assets["model"].predict_proba(df)[0]

                winner_map = {2: "Home Win", 1: "Draw", 0: "Away Win"}
                color_map = {2: "green", 1: "gray", 0: "red"}

                result = winner_map[prediction]
                # Short code for Excel
                result_code = {2: "H", 1: "D", 0: "A"}[prediction]
                
                st.markdown(f"### :{color_map[prediction]}[{result}]")

                st.caption("Confidence")
                st.progress(float(probs[2]), text=f"Home: {probs[2]:.1%}")
                st.progress(float(probs[1]), text=f"Draw: {probs[1]:.1%}")
                st.progress(float(probs[0]), text=f"Away: {probs[0]:.1%}")

                # Collect result for Excel
                current_match_results.append({
                    "Order": len(st.session_state['prediction_history']),
                    "Gameweek": gameweek,
                    "Date": match_date.strftime("%d/%m/%Y"),
                    "Time": match_time.strftime("%H:%M"),
                    "HomeTeam": home_team,
                    "AwayTeam": away_team,
                    "Model": model_name.lower().replace(" (pure 26)", ""),
                    "Result": result_code
                })

            except Exception as e:
                st.error(f"Prediction Error: {e}")

    # Add to session history
    if current_match_results:
        st.session_state['prediction_history'].extend(current_match_results)
        st.success(f"Added {len(current_match_results)} predictions to history!")

# -------------------------------------------------
# History & Export
# -------------------------------------------------
if st.session_state['prediction_history']:
    st.divider()
    st.subheader("Prediction History")
    
    history_df = pd.DataFrame(st.session_state['prediction_history'])
    
    # Pivot for storage-like display but also for Excel
    # Use Order in index to maintain sequence, then drop it for display/Excel
    pivot_df = history_df.pivot_table(
        index=["Order", "Gameweek", "Date", "Time", "HomeTeam", "AwayTeam"],
        columns="Model",
        values="Result",
        aggfunc='first'
    ).reset_index().sort_values("Order")
    
    # Drop Order for display but keep for logic if needed
    display_df = pivot_df.drop(columns="Order")
    
    st.dataframe(display_df, width='stretch')
    
    col_dl, col_clr = st.columns([1, 4])
    
    with col_dl:
        # Create Styled Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            display_df.to_excel(writer, index=False, sheet_name='Predictions')
            
            # Access the workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['Predictions']
            
            # Define Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal="center")
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_alignment

            # Apply column and row styling
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Gameweek, Date, Time (Columns 1, 2, 3)
                    if col <= 3:
                        cell.fill = orange_fill
                    # HomeTeam, AwayTeam (Columns 4, 5)
                    elif col <= 5:
                        cell.fill = yellow_fill
                    # Result columns
                    else:
                        cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        excel_data = output.getvalue()
        
        st.download_button(
            label="📥 Download Styled Excel",
            data=excel_data,
            file_name=f"Predictions_{gameweek}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    with col_clr:
        if st.button("🗑️ Clear History"):
            st.session_state['prediction_history'] = []
            st.rerun()
