import streamlit as st
import pandas as pd
import joblib
import pickle
import os
import io
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# -------------------------------------------------
# Page Config
# -------------------------------------------------
st.set_page_config(
    page_title="Premier League Predictor (No Odds)",
    layout="wide"
)

# Initialize Session State for history
if 'prediction_history' not in st.session_state:
    st.session_state['prediction_history'] = []

# -------------------------------------------------
# Models Configuration (NO XGBOOST)
# -------------------------------------------------
MODELS = {
    "Random Forest": {
        "model": "football_model_rf_no_odds.pkl",
        "encoding": "team_encoding_rf_no_odds.pkl",
        "history": "full_match_history_rf_no_odds.pkl"
    },
    "Logistic Regression": {
        "model": "football_model_no_odds.pkl",
        "encoding": "team_encoding_no_odds.pkl",
        "history": "full_match_history_no_odds.pkl"
    },
    "XGBoost (Pure 26)": {
        "model": "football_model_pure26_no_odd.pkl",
        "encoding": "team_encoding_pure26_no_odd.pkl",
        "history": "full_match_history_no_odds.pkl"
    }
}

# -------------------------------------------------
# Load Resources (Cached)
# -------------------------------------------------
@st.cache_resource(show_spinner=False)
def load_all_models():
    loaded_data = {}

    for name, config in MODELS.items():
        try:
            if not os.path.exists(config["model"]):
                st.warning(f"⚠️ {name}: Model file not found.")
                continue

            model = joblib.load(config["model"])

            with open(config["encoding"], "rb") as f:
                encoding = pickle.load(f)

            with open(config["history"], "rb") as f:
                history = pickle.load(f)

            loaded_data[name] = {
                "model": model,
                "encoding": encoding,
                "history": history
            }

        except Exception as e:
            st.error(f"❌ Error loading {name}: {e}")

    return loaded_data


model_registry = load_all_models()

# -------------------------------------------------
# Helper Functions
# -------------------------------------------------
def get_rolling_stats(team_name, df, is_home_team):
    team_matches = df[df["HomeTeam"] == team_name]
    if team_matches.empty:
        return None

    if "Date" in team_matches.columns:
        last_match = team_matches.sort_values("Date", ascending=False).iloc[0]
    else:
        last_match = team_matches.iloc[-1]

    cols = ["FTHG", "FTAG", "HS", "AS", "HST", "AST"]
    rolling_cols = [f"{c}_rolling" for c in cols]

    stats = last_match[rolling_cols].to_dict()

    if not is_home_team:
        stats = {f"Away_{k}": v for k, v in stats.items()}

    return stats

# -------------------------------------------------
# UI
# -------------------------------------------------
st.title("Premier League Predictor")
st.markdown("### Compare predictions (No Betting Odds)")

if not model_registry:
    st.error("No models loaded successfully. Please check your files.")
    st.stop()

# Team list
first_model_key = list(model_registry.keys())[0]
teams = sorted(model_registry[first_model_key]["encoding"].keys())

st.sidebar.header("Match Details")

col1, col2 = st.columns(2)
with col1:
    home_team = st.selectbox("Home Team", teams)
with col2:
    away_team = st.selectbox("Away Team", teams, index=1)

col3, col4 = st.columns(2)
with col3:
    match_date = st.date_input("Match Date")
with col4:
    match_time = st.time_input("Match Time")

st.sidebar.divider()
gameweek = st.sidebar.text_input("Gameweek", value="Gameweek17")

# -------------------------------------------------
# Prediction Logic
# -------------------------------------------------
if st.button("Run Predictions", type="primary"):
    if home_team == away_team:
        st.error("Home and Away teams cannot be the same!")
        st.stop()

    st.divider()
    cols = st.columns(len(model_registry))

    # Temporary list for current run (all models)
    current_match_results = []
    
    for col, (model_name, assets) in zip(cols, model_registry.items()):
        with col:
            st.subheader(model_name)

            day_code = match_date.weekday()
            hour = match_time.hour

            input_data = {
                "home_code": assets["encoding"].get(home_team, -1),
                "away_code": assets["encoding"].get(away_team, -1),
                "hour": hour,
                "day_code": day_code
            }

            home_stats = get_rolling_stats(home_team, assets["history"], True)
            away_stats = get_rolling_stats(away_team, assets["history"], False)

            if not home_stats or not away_stats:
                st.warning("Insufficient history data.")
                continue

            input_data.update(home_stats)
            input_data.update(away_stats)

            try:
                df = pd.DataFrame([input_data])

                prediction = assets["model"].predict(df)[0]
                probs = assets["model"].predict_proba(df)[0]

                winner_map = {2: "Home Win", 1: "Draw", 0: "Away Win"}
                color_map = {2: "green", 1: "gray", 0: "red"}

                result = winner_map[prediction]
                # Short code for Excel
                result_code = {2: "H", 1: "D", 0: "A"}[prediction]
                
                st.markdown(f"### :{color_map[prediction]}[{result}]")

                st.caption("Confidence")
                st.progress(float(probs[2]), text=f"Home: {probs[2]:.1%}")
                st.progress(float(probs[1]), text=f"Draw: {probs[1]:.1%}")
                st.progress(float(probs[0]), text=f"Away: {probs[0]:.1%}")

                # Collect result for Excel
                current_match_results.append({
                    "Gameweek": gameweek,
                    "Date": match_date.strftime("%d/%m/%Y"),
                    "Time": match_time.strftime("%H:%M"),
                    "HomeTeam": home_team,
                    "AwayTeam": away_team,
                    "Model": model_name.lower().replace(" (pure 26)", ""),
                    "Result": result_code
                })

            except Exception as e:
                st.error(f"Prediction Error: {e}")

    # Add to session history
    if current_match_results:
        st.session_state['prediction_history'].extend(current_match_results)
        st.success(f"Added {len(current_match_results)} predictions to history!")

# -------------------------------------------------
# History & Export
# -------------------------------------------------
if st.session_state['prediction_history']:
    st.divider()
    st.subheader("Prediction History")
    
    history_df = pd.DataFrame(st.session_state['prediction_history'])
    
    # Pivot for storage-like display but also for Excel
    pivot_df = history_df.pivot_table(
        index=["Gameweek", "Date", "Time", "HomeTeam", "AwayTeam"],
        columns="Model",
        values="Result",
        aggfunc='first'
    ).reset_index()
    
    st.dataframe(pivot_df, use_container_width=True)
    
    col_dl, col_clr = st.columns([1, 4])
    
    with col_dl:
        # Create Styled Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, index=False, sheet_name='Predictions')
            
            # Access the workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['Predictions']
            
            # Define Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal="center")
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_alignment

            # Apply column and row styling
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Gameweek, Date, Time (Columns 1, 2, 3)
                    if col <= 3:
                        cell.fill = orange_fill
                    # HomeTeam, AwayTeam (Columns 4, 5)
                    elif col <= 5:
                        cell.fill = yellow_fill
                    # Result columns
                    else:
                        cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        excel_data = output.getvalue()
        
        st.download_button(
            label="📥 Download Styled Excel",
            data=excel_data,
            file_name=f"Predictions_{gameweek}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    with col_clr:
        if st.button("🗑️ Clear History"):
            st.session_state['prediction_history'] = []
            st.rerun()
